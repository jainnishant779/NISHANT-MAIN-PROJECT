from openpyxl import load_workbook
from selenium import webdriver
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

driver = webdriver.Chrome(ChromeDriverManager().install())

# Load the Excel workbook
workbook = load_workbook('Book1.xlsx')
sheet = workbook.active

# Initialize Selenium WebDriver
# driver = webdriver.Chrome()


# Open the website
driver.get('https://eservices.icai.org/EForms/configuredHtml/1666/57499/Registration.html?action=newuser')

# Fill the form
for row in sheet.iter_rows(values_only=True):
    # Fill in the form fields using row data
    driver.find_element_by_id('E-Mail Address').send_keys(row[1])
    driver.find_element_by_id('Name').send_keys(row[0])
    driver.find_element_by_id('Mobile').send_keys(str(row[2]))  # Assuming Mobile is a number
    driver.find_element_by_id('Address').send_keys(row[3])
    # Continue filling other fields as needed
    
    # Submit the form
  #  driver.find_element_by_id('Submit').click()

# Close the browser
#driver.quit()
